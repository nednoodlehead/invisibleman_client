# capabilities to export data to csv / excel sheet
# uses openpxyl & csv
# examples: https://foss.heptapod.net/openpyxl/openpyxl
import os
import shutil
import openpyxl
from openpyxl.worksheet.table import TABLESTYLES, Table, TableStyleInfo
from openpyxl.styles import Color, PatternFill
import csv
import sqlite3
import datetime
from db.fetch import fetch_all
from util.data_types import InventoryObject
from db.fetch import (
    fetch_all_enabled,
    fetch_obj_from_retired,
    fetch_obj_from_loc,
    fetch_retired_assets,
    fetch_changed_assets
)
import datetime
import dateutil.relativedelta
# self is passed in so we can update the self.reports_export_status_content label
# in rust i would just pass the result<> up the stream, but that design pattern doesnt really fit
# too well in python. also passed in to see self.main_table

# im also not too sure if pulling from main_table or db is quicker, im assuming db is better..

# TODO refactor this file so its less redundant.. dont need 4 if csv statements


def export_active(self, csv: bool, file: str):
    csv_or_xlsx_str = ".csv" if csv is True else ".xlsx"
    file = f"{file}_ACTIVE{csv_or_xlsx_str}"
    data = fetch_all_enabled(self.connection)
    if csv:
        write_iter_into_csv(self.default_columns, data, file)
    else:  # excel export :D
        write_iter_into_excel(self.default_columns, data, file)
    open_explorer_at_file(self, file)


def export_retired(self, csv: bool, file: str):
    csv_or_xlsx_str = ".csv" if csv is True else ".xlsx"
    file = f"{file}_RETIRED{csv_or_xlsx_str}"
    data = fetch_obj_from_retired(self.connection)
    all_col = self.default_columns.copy()
    all_col.append("Retirement Date")
    if csv:
        write_iter_into_csv(all_col, data, file)
    else:  # excel export :D
        write_iter_into_excel(all_col, data, file)
    open_explorer_at_file(self, file)


def export_loc(self, csv: bool, file: str, place: str):
    csv_or_xlsx_str = ".csv" if csv is True else ".xlsx"
    file = f"{file}_LOCATION{csv_or_xlsx_str}"
    data = fetch_obj_from_loc(self.connection, place)
    if csv:
        write_iter_into_csv(self.default_columns, data, file)
    else:  # excel export :D
        write_iter_into_excel(self.default_columns, data, file)
    open_explorer_at_file(self, file)


def export_replacementdate(self, csv: bool, file: str, time_period: str, include_overdue):
    # time period is our 3, 6, 9, 12, 24 month period
    csv_or_xlsx_str = ".csv" if csv is True else ".xlsx"
    file = f"{file}_DUE-REPLACEMENT{csv_or_xlsx_str}"
    num = int(time_period.split(" ")[0])  # the number part of the '3 Months' or '12 Months'
    time_period = datetime.date.today() + dateutil.relativedelta.relativedelta(months=num)
    data = fetch_retired_assets(self.connection, time_period, include_overdue)
    if csv:
        write_iter_into_csv(self.default_columns, data, file)
    else:  # excel export :D
        write_iter_into_excel(self.default_columns, data, file)
    open_explorer_at_file(self, file)


def write_iter_into_csv(columns, iterable, filename: str):
    with open(filename, "w") as csv_file:
        wr = csv.writer(csv_file, delimiter=",")
        wr.writerow(columns)
        for item in iterable:
            wr.writerow(list(item))


def write_iter_into_excel(columns, iterable, filename: str):
    wb = openpyxl.Workbook()
    letter_list = ["a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m"]
    row_hex = "9A9C9B"
    active_ws = wb.active
    color = PatternFill(start_color=row_hex, end_color=row_hex, fill_type='solid')
    active_ws.append(columns)
    for count, row in enumerate(iterable):
        # so the count must begin at 1, but if the 
        active_ws.append(row)
        if (count + 1) % 2 == 0:
            for internal_count, _ in enumerate(columns):
                print("painting", letter_list[internal_count], count)
                active_ws[f'{letter_list[internal_count]}{count}'].fill = color                
    for count, col in enumerate(letter_list):
        active_ws.column_dimensions[col].width = 20
    # we add +1 because indexes start at 0 in py, but start at one in excel..
    table = Table(displayName="exported", ref=f'A1:{letter_list[len(columns) -1]}{len(iterable) +1}')
    active_ws.add_table(table)
    wb.save(filename)


# not really an 'export' file, but is only used here, an fits nowhere else really..
def open_explorer_at_file(self, file: str):
    # maybe have setting (only one or the other) to open the file in explorer, or to open the file..
    file = os.path.dirname(file)

    if self.settings_report_auto_open_checkbox.isChecked():
        file = os.path.realpath(file)
        os.startfile(file)


def create_backup(self):
    time_and_date = str(datetime.datetime.now()).replace(":", "-")[:19]
    filename = f"INVMAN_BACKUP__{time_and_date}"
    backup_dir = self.config["backup_path"]
    if not backup_dir.endswith("/") or not backup_dir.endswith("\\"):
        backup_dir += "/"
    filename += ".db"
    end_file = f"{backup_dir}{filename}"
    shutil.copyfile("main.db", end_file)
    self.display_message("Success!", f"Backup created successfully: {filename}")
    open_explorer_at_file(self, end_file)

def export_changed(self, csv: bool, file: str, month, year):
    # we are going to export the entire `changed` table
    csv_or_xlsx_str = ".csv" if csv is True else ".xlsx"
    data = fetch_changed_assets(self.connection, month, year)
    special_columns = ["Old Name", "New Name", "Old Location", "New Location", "Date Changed"]
    file = f'{file}_CHANGED{csv_or_xlsx_str}'
    if csv:
        write_iter_into_csv(special_columns, data, file)
    else:  # excel export :D
        write_iter_into_excel(special_columns, data, file)
    open_explorer_at_file(self, file)

def compare_intune_and_invisman(self):
    intune_file = self.reports_utilities_intune_file_path.text()
    if intune_file is None:
        self.display_message("Error!", "You do not have a selected path in the 'Intune Comparison' text box")
    # this file is provided by intune
    # you can find it from the Itune admin center -> Devices -> Windows Devices -> "Export" button
    # it will create a .zip, and inside that .zip is an excel file with the data
    # much of the data is useless, but it does provided a useful amount of extra stuff
    # this is primar comparing for deficiencies (devices exists in one but not the other)
    # it also looks for mismatches in serial number and names
    # it will also do checks to ensure that column names are where they should be
    class IntuneObj:
        def __init__(self, device_name, sn, manu, model):
            self.name = device_name
            self.serial = sn
            self.manufacturer = manu
            self.model = model
        
        def __repr__(self):
            return f'{self.name} | {self.serial}'

        def __iter__(self):
            yield self.name
            yield self.serial
            yield self.manufacturer
            yield self.model
    intune_list = []
    with open(intune_file, "r") as file:
        reader = csv.reader(file)
        for count, row in enumerate(reader):
            if count == 0:
                # we do a quick check to ensure microsoft hasnt done something stupid and changed the column exports for some reason
                if row[1] != "Device name":
                    self.display_message("Intune Error", "The column name at position 'B' is not 'Device name'")
                if row[8] != "Serial number":
                    self.display_message("Intune Error", "The column name at letter 'I' is not 'Serial number'")
                if row[9] != "Manufacturer":
                    self.display_message("Intune Error", "The column name at letter 'J' is not 'Manufacturer'")
                if row[10] != "Model":
                    self.display_message("Intune Error", "The column name at letter 'K' is not 'Model'")
                continue
            intune_list.append(IntuneObj(row[1], row[8], row[9], row[10]))
    invisman_db = fetch_all(self.connection)
    in_invis_not_intune = []
    in_intune_not_invis = []
    discrepancy = []
    both = []
    for invent in invisman_db:
        for intune in intune_list:
            if invent.serial == intune.serial:
                both.append(invent.serial)
                if invent.name != intune.name:
                    discrepancy.append((invent, intune))
    for sn in invisman_db:
        if sn.serial not in both:
            in_invis_not_intune.append(sn)
    for sn in intune_list:
        if sn.serial not in both:
            in_intune_not_invis.append(sn)
    export_file = self.export_file_path_choice.text()
    write_comparison_to_excel(self, export_file, in_invis_not_intune, in_intune_not_invis, discrepancy)
    print(in_invis_not_intune, "\n", "\n")
    print(in_intune_not_invis, "\n")
    for (inv, intu) in discrepancy:
        print(f'inv: {inv} | intune: {intu} ')
    print(len(both))

def write_comparison_to_excel(self, export_dir, invis_only, intune_only, discrepancy):
    # honestly, writing this into csv seems sort of trivial and not very useful
    # well, not trivial because it's hard, but trivial because there is no great spot for the radio button..
    # also how would you represent the 3 different types in a csv file?
    # in invisman, not intune. in intune, not invisman, discrepancies in naming
    full_file_path = f'{export_dir}/intune-comparison-{str(datetime.datetime.now()).replace(":", "-")[:19]}.xlsx'
    # does "export_dir" already have a slash at the end???
    workb = openpyxl.Workbook()
    ws = workb.active
    ws["A1"] = "Exists in invisman ✅ Exists in intune ❌"
    ws["A2"] = "Name"
    ws["B2"] = "Serial"
    ws["C2"] = "Manufacturer"
    ws["D2"] = "Model"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 30
    for count, invis_obj in enumerate(invis_only):
        ws[f'A{count +3}'] = invis_obj.name
        ws[f'B{count +3}'] = invis_obj.serial
        ws[f'C{count +3}'] = invis_obj.manufacturer
        ws[f'D{count +3}'] = invis_obj.model
    sty = TableStyleInfo(name="TableStyleDark1", showRowStripes=True, showFirstColumn=False, showLastColumn=False, showColumnStripes=False)
    table = Table(displayName="Exported", ref=f'A2:D{count+3}')
    table.tableStyleInfo = sty
    ws.add_table(table)
    start_second_table = f'A{count +5}'
    ws[f"A{count+ 4}"] = "Exists in invisman ✅ Exists in intune ❌"
    ws[f"A{count+ 5}"] = "Name"
    ws[f"B{count+ 5}"] = "Serial"
    ws[f"C{count+ 5}"] = "Manufacturer"
    ws[f"D{count+ 5}"] = "Model"
    
    for second_count, intune_obj in enumerate(intune_only):
        # +6 because index starts at 1 (+3) and the +3 from the stuff above ^^ and +2 from the titles for this row
        ws[f'A{count + second_count +6}'] = intune_obj.name
        ws[f'B{count + second_count +6}'] = intune_obj.serial
        ws[f'C{count + second_count +6}'] = intune_obj.manufacturer
        ws[f'D{count + second_count +6}'] = intune_obj.model

    sty = TableStyleInfo(name="TableStyleDark3", showRowStripes=True, showFirstColumn=False, showLastColumn=False, showColumnStripes=False)
    table_2 = Table(displayName="da_second", ref=f'{start_second_table}:D{second_count + count +6}')
    table_2.tableStyleInfo = sty
    ws.add_table(table_2)
    third_table_begin = f'A{count + second_count + 8}'
    ws[f'A{count + second_count + 7}'] = "Data Mismatch❌❌"
    ws[f"A{count + second_count + 8}"] = "Name (intune)"
    ws[f"B{count + second_count + 8}"] = "Serial"
    ws[f"C{count + second_count + 8}"] = "Manufacturer"
    ws[f"D{count + second_count + 8}"] = "Model"
    ws[f"E{count + second_count + 8}"] = "Name (INVISMAN)"
    ws[f"F{count + second_count + 8}"] = "Serial (Matching)"
    ws[f"G{count + second_count + 8}"] = "Manufacturer (May match)"
    ws[f"H{count + second_count + 8}"] = "Model (Should Match)"
    for third_count, (intune_obj, invis_obj) in enumerate(discrepancy):
        ws[f'A{count + second_count + third_count +9}'] = intune_obj.name
        ws[f'B{count + second_count + third_count +9}'] = intune_obj.serial
        ws[f'C{count + second_count + third_count +9}'] = intune_obj.manufacturer
        ws[f'D{count + second_count + third_count +9}'] = intune_obj.model
        ws[f'E{count + second_count + third_count +9}'] = invis_obj.name
        ws[f'F{count + second_count + third_count +9}'] = invis_obj.serial
        ws[f'G{count + second_count + third_count +9}'] = invis_obj.manufacturer
        ws[f'H{count + second_count + third_count +9}'] = invis_obj.model
    
    sty = TableStyleInfo(name="TableStyleDark5", showRowStripes=True, showFirstColumn=False, showLastColumn=False, showColumnStripes=False)
    table_3 = Table(displayName="da_third", ref=f'{third_table_begin}:H{third_count + second_count + count +9}')
    table_3.tableStyleInfo = sty
    ws.add_table(table_3)



        
    workb.save(full_file_path)
    open_explorer_at_file(self, full_file_path)
    

def compare_ip_to_site_and_invisman(self):
    pass
